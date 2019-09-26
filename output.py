import csv
import openpyxl
from openpyxl import Workbook

def WriteToCSV(filename, articles):
    with open(filename, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        for article in articles:
            writer.writerow([article.title, article.url])

def WriteToXLS(filename, articles):
    wb = Workbook()
    ws = wb.active
    # format spreadsheet columns
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 60
    for i, article in enumerate(articles, start=1):
        title_cell = ws.cell(row = i, column = 1, value = article.title)
        url_cell = ws.cell(row = i, column = 2, value = article.url)
        url_cell.hyperlink = article.url
        url_cell.style = 'Hyperlink'
        date_cell = ws.cell(row = i, column = 3, value = article.date)
        snippet_cell = ws.cell(row = i, column = 4, value = article.snippet)
    wb.save(filename)
